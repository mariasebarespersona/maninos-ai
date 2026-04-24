"""
Import QuickBooks Chart of Accounts from Excel export into accounting_accounts table.

Usage: python scripts/import_quickbooks_accounts.py

This script:
1. Reads the QuickBooks export Excel file
2. Parses the hierarchy from "Full name" column (colon-separated)
3. Creates all accounts preserving parent-child relationships
4. Uses QuickBooks account types directly (no mapping)
"""

import os
import sys
import openpyxl
from supabase import create_client

# Supabase connection
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://tpsszoxyqdutqlwfgrvm.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_KEY:
    print("ERROR: Set SUPABASE_SERVICE_ROLE_KEY env var")
    sys.exit(1)

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

EXCEL_PATH = "/Users/mariasebares/Downloads/Report 04_16_2026T10_30_36.xlsx"

# P&L types (for report filtering)
PL_TYPES = {"Income", "Other Income", "Cost of Goods Sold", "Expenses", "Other Expense"}
# BS types
BS_TYPES = {"Bank", "Accounts receivable (A/R)", "Other Current Assets", "Fixed Assets",
            "Other Assets", "Accounts payable (A/P)", "Other Current Liabilities",
            "Long Term Liabilities", "Equity"}


def parse_excel():
    """Parse Excel and return list of account dicts."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    accounts = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        full_name = row[0]
        acct_type = row[1]
        detail_type = row[2]
        description = row[3]
        balance = row[4]
        short_name = row[8]  # Column I = "Account" = short display name

        if not full_name or not acct_type:
            continue

        full_name = str(full_name).strip()
        acct_type = str(acct_type).strip()
        detail_type = str(detail_type).strip() if detail_type else ""
        description = str(description).strip() if description else ""
        short_name = str(short_name).strip() if short_name else full_name.split(":")[-1].strip()

        # Parse hierarchy from full_name (colon-separated)
        parts = [p.strip() for p in full_name.split(":")]
        depth = len(parts) - 1
        parent_path = ":".join(parts[:-1]) if depth > 0 else None

        # Balance
        try:
            bal = float(balance) if balance else 0.0
        except (ValueError, TypeError):
            bal = 0.0

        accounts.append({
            "full_name": full_name,
            "short_name": short_name,
            "account_type": acct_type,
            "detail_type": detail_type,
            "description": description,
            "balance": bal,
            "depth": depth,
            "parent_path": parent_path,
        })

    return accounts


def import_accounts(accounts):
    """Import accounts into Supabase, preserving hierarchy."""

    # Step 1: Delete all existing accounts
    print("Deleting existing accounts...")
    # First clear FK refs
    sb.table("accounting_transactions").update({"account_id": None}).neq("id", "00000000-0000-0000-0000-000000000000").execute()
    sb.table("statement_movements").update({"suggested_account_id": None, "final_account_id": None}).neq("id", "00000000-0000-0000-0000-000000000000").execute()

    # Delete in order (children first)
    existing = sb.table("accounting_accounts").select("id, parent_account_id").execute().data or []
    # Delete children first
    children = [a for a in existing if a.get("parent_account_id")]
    parents = [a for a in existing if not a.get("parent_account_id")]
    for a in children:
        try:
            sb.table("accounting_accounts").delete().eq("id", a["id"]).execute()
        except Exception:
            pass
    for a in parents:
        try:
            sb.table("accounting_accounts").delete().eq("id", a["id"]).execute()
        except Exception:
            pass

    # Verify clean
    remaining = sb.table("accounting_accounts").select("id", count="exact").execute()
    print(f"  Remaining after delete: {len(remaining.data or [])}")

    # Step 2: Create root section headers (P&L and BS structure)
    print("\nCreating section headers...")
    sections = [
        # P&L sections
        {"code": "PL_INCOME", "name": "Income", "account_type": "Income", "is_header": True, "display_order": 1000},
        {"code": "PL_OTHER_INCOME", "name": "Other Income", "account_type": "Other Income", "is_header": True, "display_order": 1500},
        {"code": "PL_COGS", "name": "Cost of Goods Sold", "account_type": "Cost of Goods Sold", "is_header": True, "display_order": 2000},
        {"code": "PL_EXPENSES", "name": "Expenses", "account_type": "Expenses", "is_header": True, "display_order": 3000},
        {"code": "PL_OTHER_EXPENSES", "name": "Other Expense", "account_type": "Other Expense", "is_header": True, "display_order": 3500},
        # BS sections
        {"code": "BS_ASSETS", "name": "Assets", "account_type": "Bank", "is_header": True, "display_order": 4000},
        {"code": "BS_LIABILITIES", "name": "Liabilities", "account_type": "Accounts payable (A/P)", "is_header": True, "display_order": 5000},
        {"code": "BS_EQUITY", "name": "Equity", "account_type": "Equity", "is_header": True, "display_order": 6000},
    ]

    section_ids = {}
    for s in sections:
        result = sb.table("accounting_accounts").insert(s).execute()
        if result.data:
            section_ids[s["code"]] = result.data[0]["id"]
            print(f"  Created {s['code']}: {s['name']}")

    # Map QB type → parent section
    type_to_section = {
        "Income": "PL_INCOME",
        "Other Income": "PL_OTHER_INCOME",
        "Cost of Goods Sold": "PL_COGS",
        "Expenses": "PL_EXPENSES",
        "Other Expense": "PL_OTHER_EXPENSES",
        "Bank": "BS_ASSETS",
        "Accounts receivable (A/R)": "BS_ASSETS",
        "Other Current Assets": "BS_ASSETS",
        "Fixed Assets": "BS_ASSETS",
        "Other Assets": "BS_ASSETS",
        "Accounts payable (A/P)": "BS_LIABILITIES",
        "Other Current Liabilities": "BS_LIABILITIES",
        "Long Term Liabilities": "BS_LIABILITIES",
        "Equity": "BS_EQUITY",
    }

    # Step 3: Import accounts by depth (parents first)
    print(f"\nImporting {len(accounts)} accounts...")

    # Sort by depth so parents are created before children
    accounts.sort(key=lambda a: a["depth"])

    # Map full_name → id for parent lookups
    name_to_id = {}
    created = 0
    skipped = 0
    errors = 0

    for i, acct in enumerate(accounts):
        # Determine parent_account_id
        parent_id = None
        if acct["parent_path"]:
            parent_id = name_to_id.get(acct["parent_path"])
            if not parent_id:
                # Parent not found — use section header as parent
                section_code = type_to_section.get(acct["account_type"])
                parent_id = section_ids.get(section_code) if section_code else None
        else:
            # Top-level account — parent is the section header
            section_code = type_to_section.get(acct["account_type"])
            parent_id = section_ids.get(section_code) if section_code else None

        # Build code from short_name (clean it up)
        code = acct["short_name"].replace(" ", "_").replace("/", "-")[:50]

        # Determine if this is a header (has children)
        is_header = any(
            other["parent_path"] == acct["full_name"]
            for other in accounts
        )

        insert_data = {
            "code": code,
            "name": acct["short_name"],
            "account_type": acct["account_type"],
            "category": acct["detail_type"],
            "description": acct["description"][:500] if acct["description"] else None,
            "parent_account_id": parent_id,
            "is_header": is_header,
            "is_active": True,
            "current_balance": acct["balance"] if acct["balance"] != 0 else None,
            "display_order": i + 100,
        }
        insert_data = {k: v for k, v in insert_data.items() if v is not None}

        try:
            result = sb.table("accounting_accounts").insert(insert_data).execute()
            if result.data:
                name_to_id[acct["full_name"]] = result.data[0]["id"]
                created += 1
            else:
                skipped += 1
        except Exception as e:
            errors += 1
            if errors <= 10:
                print(f"  ERROR [{acct['short_name']}]: {str(e)[:100]}")

        if (i + 1) % 100 == 0:
            print(f"  Progress: {i+1}/{len(accounts)} (created: {created}, errors: {errors})")

    print(f"\nDone: {created} created, {skipped} skipped, {errors} errors")
    return created


if __name__ == "__main__":
    print("Parsing Excel...")
    accounts = parse_excel()
    print(f"Found {len(accounts)} accounts")

    # Show summary
    types = {}
    for a in accounts:
        t = a["account_type"]
        types[t] = types.get(t, 0) + 1
    print("\nBy type:")
    for t, c in sorted(types.items()):
        print(f"  {t}: {c}")

    print("\n" + "="*50)
    confirm = input("Import all accounts? (yes/no): ")
    if confirm.lower() != "yes":
        print("Cancelled")
        sys.exit(0)

    import_accounts(accounts)
