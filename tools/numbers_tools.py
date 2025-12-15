from __future__ import annotations
from typing import Dict, List, Optional, Union
from .supabase_client import sb
from .utils import nums_schema
from tools.formula_calculator_v3_simple import auto_calculate_on_update
from .verifier import verify_numbers_update
import logging

logger = logging.getLogger(__name__)

# ==================== MANINOS AI LOGIC ====================

# Defect costs mapping (Simple dictionary as requested)
DEFECT_COSTS = {
    'roof': 3000,
    'hvac': 2500,
    'plumbing': 1500,
    'electrical': 2000,
    'flooring': 1200,
    'windows': 1000,
    'skirting': 800,
    'painting': 1000,
    'appliances': 1500,
    'deck': 1000,
    'other': 500  # Generic cost for unspecified defects
}

def calculate_repair_costs(defects: List[str]) -> Dict:
    """
    Calculate estimated repair costs based on a list of defects.
    Returns the total cost and a breakdown.
    """
    total_cost = 0
    breakdown = {}
    
    for defect in defects:
        defect_lower = defect.lower().strip()
        # Find best match in DEFECT_COSTS
        cost = 0
        matched_key = "other"
        
        if defect_lower in DEFECT_COSTS:
            cost = DEFECT_COSTS[defect_lower]
            matched_key = defect_lower
        else:
            # Fuzzy / partial match
            found = False
            for key, val in DEFECT_COSTS.items():
                if key in defect_lower:
                    cost = val
                    matched_key = key
                    found = True
                    break
            if not found:
                cost = DEFECT_COSTS['other']
                matched_key = "other (default)"
        
        total_cost += cost
        breakdown[defect] = cost
        
    return {
        "total_cost": total_cost,
        "breakdown": breakdown
    }

def calculate_maninos_deal(
    asking_price: float, 
    repair_costs: Optional[float] = 0, 
    arv: Optional[float] = None, 
    market_value: Optional[float] = None,
    property_id: Optional[str] = None
) -> Dict:
    """
    Calculate deal viability based on Maninos AI rules.
    
    Rules:
    1. 70% Rule (Soft Filter): Asking Price <= Market Value * 0.70
       - Only checked if Market Value is provided.
       - If PASSED and property_id provided: Updates acquisition_stage to 'passed_70_rule'
    2. 80% Rule (Hard Filter): (Asking Price + Repair Costs) <= (ARV * 0.80)
       - Only checked if ARV is provided.
       - If PASSED and property_id provided: Updates acquisition_stage to 'passed_80_rule'
       - If FAILED and property_id provided: Updates acquisition_stage to 'rejected'
    
    This function can be called at different stages:
    - Step 1: With only asking_price and market_value (checks 70% rule only)
    - Step 4: With all parameters (checks both 70% and 80% rules)
    
    Args:
        asking_price: Purchase price
        repair_costs: Estimated repair costs
        arv: After Repair Value
        market_value: Current market value
        property_id: Optional UUID to update acquisition_stage in DB
    
    Returns:
        Dict with status, reasoning, metrics, and updated acquisition_stage
    """
    # Initialize result
    result = {
        "status": "Review Required", # Default
        "metrics": {
            "asking_price": asking_price,
            "repair_costs": repair_costs or 0,
            "arv": arv,
            "market_value": market_value,
            "total_investment": asking_price + (repair_costs or 0),
            "max_allowable_offer_70": 0,
            "max_investment_80": (arv * 0.80) if arv else None
        },
        "checks": {
            "70_percent_rule": None,
            "80_percent_rule": None
        },
        "reasoning": []
    }
    
    # 1. Check 70% Rule (if market_value is present)
    if market_value:
        max_offer_70 = market_value * 0.70
        result["metrics"]["max_allowable_offer_70"] = max_offer_70
        
        if asking_price <= max_offer_70:
            result["checks"]["70_percent_rule"] = "PASS"
            result["reasoning"].append(f"‚úÖ 70% Rule PASS: Asking Price (${asking_price:,.0f}) is within 70% of Market Value (max: ${max_offer_70:,.0f}).")
        else:
            result["checks"]["70_percent_rule"] = "FAIL"
            result["reasoning"].append(f"‚ö†Ô∏è 70% Rule WARNING: Asking Price (${asking_price:,.0f}) exceeds 70% of Market Value (max: ${max_offer_70:,.0f}).")
            result["status"] = "Review Required"
            
    # 2. Check 80% Rule (Hard Filter) - only if ARV is provided
    if arv:
        total_investment = asking_price + (repair_costs or 0)
        limit_80 = arv * 0.80
        
        if total_investment <= limit_80:
            result["checks"]["80_percent_rule"] = "PASS"
            # Only mark as "Ready to Buy" if BOTH rules pass (or 70% rule wasn't checked)
            if result["checks"]["70_percent_rule"] in ["PASS", None]:
                result["status"] = "Ready to Buy"
            result["reasoning"].append(f"‚úÖ 80% Rule PASS: Total Investment (${total_investment:,.0f}) is within 80% of ARV (max: ${limit_80:,.0f}).")
        else:
            result["checks"]["80_percent_rule"] = "FAIL"
            result["status"] = "Rejected"
            result["reasoning"].append(f"üî¥ 80% Rule FAIL: Total Investment (${total_investment:,.0f}) exceeds 80% of ARV (max: ${limit_80:,.0f}).")
    
    # If only 70% rule was checked (Step 1), set appropriate status
    if market_value and not arv:
        if result["checks"]["70_percent_rule"] == "PASS":
            result["status"] = "Proceed to Inspection"
            result["reasoning"].append("Next: Proceed to Step 2 (Inspection & Data Collection).")
        else:
            result["status"] = "Review Required"
            result["reasoning"].append("User may proceed with justification, or reject the deal.")
    
    # ==================== UPDATE ACQUISITION STAGE (MANINOS AI) ====================
    if property_id:
        from .property_tools import update_acquisition_stage, update_property_fields
        
        try:
            # 1. SAVE PROPERTY DATA (asking_price, market_value, arv, status)
            fields_to_update = {}
            if asking_price is not None:
                fields_to_update["asking_price"] = asking_price
            if market_value is not None:
                fields_to_update["market_value"] = market_value
            if arv is not None:
                fields_to_update["arv"] = arv
            
            # Also update the status column based on result
            fields_to_update["status"] = result["status"]
            
            if fields_to_update:
                update_result = update_property_fields(property_id, fields_to_update)
                if update_result.get("ok"):
                    logger.info(f"‚úÖ [calculate_maninos_deal] Saved property data: {list(fields_to_update.keys())}")
                else:
                    logger.error(f"‚ùå [calculate_maninos_deal] Failed to save property data: {update_result.get('error')}")
            
            # 2. UPDATE ACQUISITION STAGE
            new_stage = None
            
            # If 70% rule was checked (Step 1 - Initial Submission)
            if result["checks"]["70_percent_rule"] is not None and result["checks"]["80_percent_rule"] is None:
                if result["checks"]["70_percent_rule"] == "PASS":
                    # 70% rule passed ‚Üí can proceed to inspection
                    new_stage = "passed_70_rule"
                    logger.info(f"[calculate_maninos_deal] ‚úÖ Step 1: 70% rule PASSED ‚Üí updating stage to 'passed_70_rule'")
                else:
                    # 70% rule failed ‚Üí requires human justification before proceeding
                    new_stage = "review_required"
                    logger.info(f"[calculate_maninos_deal] ‚ö†Ô∏è Step 1: 70% rule FAILED ‚Üí updating stage to 'review_required' (BLOCKED until justification)")
            
            # If 80% rule passed (Step 4)
            elif result["checks"]["80_percent_rule"] == "PASS":
                new_stage = "passed_80_rule"
                logger.info(f"[calculate_maninos_deal] 80% rule passed ‚Üí updating stage to 'passed_80_rule'")
            
            # If 80% rule failed (Step 4)
            elif result["checks"]["80_percent_rule"] == "FAIL":
                new_stage = "rejected"
                logger.info(f"[calculate_maninos_deal] 80% rule failed ‚Üí updating stage to 'rejected'")
            
            # Update stage if determined
            if new_stage:
                stage_result = update_acquisition_stage(property_id, new_stage)
                if stage_result.get("ok"):
                    result["acquisition_stage_updated"] = new_stage
                    logger.info(f"‚úÖ [calculate_maninos_deal] Acquisition stage updated: {new_stage}")
                else:
                    logger.error(f"‚ùå [calculate_maninos_deal] Failed to update stage: {stage_result.get('error')}")
        
        except Exception as e:
            logger.error(f"‚ùå [calculate_maninos_deal] Error updating property: {e}")
            # Don't fail the calculation if updates fail
    
    return result

# ==================== EXISTING CODE BELOW ====================

# R2B Template Formulas (from docs/R2B_FORMULAS.md)
# These formulas are automatically injected when importing an R2B Excel file
# All formulas use IF() to show blank cells instead of 0 when inputs are empty
R2B_FORMULAS = {
    # IVA Calculations (Column D) - Show blank if B cell is empty
    "D5": '=IF(OR(B5="",C5=""),"",B5*C5/100)',
    "D6": '=IF(OR(B6="",C6=""),"",B6*C6/100)',
    "D7": '=IF(OR(B7="",C7=""),"",B7*C7/100)',
    "D8": '=IF(OR(B8="",C8=""),"",B8*C8/100)',
    
    # Total with VAT (Column E) - Show blank if B cell is empty
    "E5": '=IF(B5="","",B5+D5)',
    "E6": '=IF(B6="","",B6+D6)',
    "E7": '=IF(B7="","",B7+D7)',
    "E8": '=IF(B8="","",B8+D8)',
    
    # Profit Calculations - Show blank if required inputs are empty
    "B10": '=IF(B6="","",B6-B7-B8)',      # Gross profit from land sale
    "B12": '=IF(B10="","",B10+B11)',      # Total gross income
    "B13": '=IF(B12="","",B12*0.25)',     # Taxes at 25%
    "B14": '=IF(B13="","",B13)',          # Taxes in euros
    "B15": '=IF(B12="","",B12-B14)',      # Net profit
    
    # AUTOPROMOCI√ìN
    "B18": '=IF(B15="","",B15)',          # Reference to net profit
    
    # Coste Comprador Total - Show blank if all inputs are empty
    "B29": '=IF(AND(B25="",B26="",B27="",B28=""),"",B25+B26+B27+B28)'  # Sum of all buyer costs
}

def set_number(property_id: str, item_key: str, amount: Optional[float]) -> Dict:
    """Set a numeric input in the numbers framework. Returns validated result."""
    import logging
    logger = logging.getLogger(__name__)
    
    schema = nums_schema(property_id)
    try:
        sb.postgrest.schema = schema
        result = (sb.table("line_items")
          .update({"amount": amount})
          .eq("property_id", property_id)
          .eq("item_key", item_key)
          .execute())
        
        # Validate that the value was saved correctly
        if result.data:
            # Verify by reading back the value
            verify = (sb.table("line_items")
              .select("amount")
              .eq("property_id", property_id)
              .eq("item_key", item_key)
              .execute())
            if verify.data and len(verify.data) > 0:
                saved_amount = verify.data[0].get("amount")
                if saved_amount == amount or (saved_amount is None and amount is None):
                    logger.info(f"‚úÖ Validated: {item_key} = {amount} saved correctly")
                    return {"ok": True, "item_key": item_key, "amount": amount, "validated": True}
                else:
                    logger.warning(f"‚ö†Ô∏è Validation failed: expected {amount}, got {saved_amount}")
                    return {"ok": False, "item_key": item_key, "error": f"Validation failed: expected {amount}, got {saved_amount}"}
        
        return {"ok": True, "item_key": item_key, "amount": amount, "validated": False}
    except Exception as e:
        logger.error(f"Error setting number {item_key}: {e}")
        # Fallback via RPC in public schema
        try:
            sb.postgrest.schema = "public"
            sb.rpc("set_property_number", {"p_id": property_id, "k": item_key, "amount": amount}).execute()
            # Try to validate by reading back from line_items
            try:
                verify = (sb.table("line_items")
                  .select("amount")
                  .eq("property_id", property_id)
                  .eq("item_key", item_key)
                  .execute())
                if verify.data and len(verify.data) > 0:
                    saved_amount = verify.data[0].get("amount")
                    if saved_amount == amount or (saved_amount is None and amount is None):
                        return {"ok": True, "item_key": item_key, "amount": amount, "validated": True}
            except:
                pass  # If validation fails, still return success but unvalidated
            return {"ok": True, "item_key": item_key, "amount": amount, "validated": False}
        except Exception as e2:
            logger.error(f"Fallback RPC also failed: {e2}")
            return {"ok": False, "item_key": item_key, "error": str(e2)}

def get_numbers(property_id: str, template_key: Optional[str] = None) -> List[Dict]:
    """Get all numbers for a property. Returns the structure even if values are NULL.
    
    If no items are found in the DB, returns the template structure for R2B by default.
    """
    # Define the template structures first
    template_structures = {
        "R2B": [
            {"group_name": "B¬∫ RAMA", "item_key": "precio_venta", "item_label": "Precio de venta", "is_percent": False, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_urbano", "item_label": "Terreno urbano", "is_percent": False, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_rustico", "item_label": "Terreno r√∫stico", "is_percent": False, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_urbano_iva_pct", "item_label": "IVA (%)", "is_percent": True, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_rustico_iva_pct", "item_label": "IVA (%)", "is_percent": True, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "project_mgmt_fees", "item_label": "Project Mgmt fees", "is_percent": False, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "impuestos_pct", "item_label": "Impuestos (%)", "is_percent": True, "amount": None},
            {"group_name": "B¬∫ RAMA", "item_key": "total_pagado", "item_label": "Total pagado a 29 julio 2025", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "terrenos_coste", "item_label": "Terrenos", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "project_management_coste", "item_label": "Project Management", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "acometidas", "item_label": "Acometidas", "is_percent": False, "amount": None},
            {"group_name": "Coste comprador", "item_key": "costes_construccion", "item_label": "Costes de construcci√≥n", "is_percent": False, "amount": None},
        ],
    }
    
    # If a template_key is explicitly provided, return the template structure merged with DB values.
    # If no template_key is provided, do NOT fall back to a default template ‚Äî return only DB items (or empty list).
    # This avoids auto-showing the old R2B template when the user didn't select it.
    if template_key:
        # Try to get items from DB and merge with template structure
        schema = nums_schema(property_id)
        db_items = {}
        try:
            sb.postgrest.schema = schema
            items = (sb.table("line_items")
                     .select("group_name,item_key,item_label,is_percent,amount,updated_at")
                     .eq("property_id", property_id)
                     .execute()).data
            # Build a map of item_key -> item for quick lookup
            for item in (items or []):
                db_items[item.get("item_key")] = item
        except Exception as e:
            # Log error for debugging
            import logging
            logging.warning(f"Error fetching from schema {schema}: {e}")
        
        # Fallback via RPC in public schema
        if not db_items:
            try:
                sb.postgrest.schema = "public"
                items = sb.rpc("list_property_numbers", {"p_id": property_id}).execute().data
                for item in (items or []):
                    db_items[item.get("item_key")] = item
            except Exception as e:
                # Log error for debugging
                import logging
                logging.warning(f"Error fetching via RPC: {e}")
        
        # Get template structure and merge with DB values
        template_structure = template_structures.get(template_key, template_structures["R2B"])
        result = []
        for template_item in template_structure:
            item_key = template_item["item_key"]
            # If we have a DB value for this item, use it (but keep template structure)
            if item_key in db_items:
                db_item = db_items[item_key]
                # Merge: use template structure but DB values for amount
                merged_item = template_item.copy()
                merged_item["amount"] = db_item.get("amount")
                merged_item["updated_at"] = db_item.get("updated_at")
                result.append(merged_item)
            else:
                # Use template structure with NULL values
                result.append(template_item.copy())
        return result
    
    # If no template_key, try to get items from DB (legacy behavior)
    schema = nums_schema(property_id)
    try:
        sb.postgrest.schema = schema
        items = (sb.table("line_items")
                 .select("item_key,amount")  # Optimization: fetch only key and amount
                 .eq("property_id", property_id)
                 .execute()).data
        if items and len(items) > 0:
            # Optimization: Return simple key:value map instead of heavy list of dicts
            return {i["item_key"]: i["amount"] for i in items}
    except Exception as e:
        import logging
        logging.warning(f"Error fetching from schema {schema}: {e}")
    
    # Fallback via RPC
    try:
        sb.postgrest.schema = "public"
        items = sb.rpc("list_property_numbers", {"p_id": property_id}).execute().data
        if items and len(items) > 0:
            # Optimization: Return simple key:value map
            return {i["item_key"]: i["amount"] for i in items}
    except Exception as e:
        import logging
        logging.warning(f"Error fetching via RPC: {e}")
    
    # If no items found in DB, return empty dict
    return {}

def calc_numbers(property_id: str) -> List[Dict]:
    schema = nums_schema(property_id)
    try:
        # This may fail if PostgREST doesn't expose the dynamic schema; try public RPC instead.
        return sb.rpc(f"{schema}.calc").execute().data
    except Exception:
        sb.postgrest.schema = "public"
        return sb.rpc("calc_property_numbers", {"p_id": property_id}).execute().data

def clear_number(property_id: str, item_key: str) -> Dict:
    """Clear/reset a specific number value by setting it to None."""
    return set_number(property_id, item_key, None)

def initialize_template_structure(property_id: str, template_key: str) -> Dict:
    """Initialize the structure of line_items for a numbers template.
    This ensures that get_numbers returns the full structure even if values are NULL.
    """
    # Define the structure for each template
    template_structures = {
        "R2B": [
            # B¬∫ RAMA section
            {"group_name": "B¬∫ RAMA", "item_key": "precio_venta", "item_label": "Precio de venta", "is_percent": False},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_urbano", "item_label": "Terreno urbano", "is_percent": False},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_rustico", "item_label": "Terreno r√∫stico", "is_percent": False},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_urbano_iva_pct", "item_label": "IVA (%)", "is_percent": True},
            {"group_name": "B¬∫ RAMA", "item_key": "terreno_rustico_iva_pct", "item_label": "IVA (%)", "is_percent": True},
            {"group_name": "B¬∫ RAMA", "item_key": "project_mgmt_fees", "item_label": "Project Mgmt fees", "is_percent": False},
            {"group_name": "B¬∫ RAMA", "item_key": "impuestos_pct", "item_label": "Impuestos (%)", "is_percent": True},
            {"group_name": "B¬∫ RAMA", "item_key": "total_pagado", "item_label": "Total pagado a 29 julio 2025", "is_percent": False},
            # Coste comprador section
            {"group_name": "Coste comprador", "item_key": "terrenos_coste", "item_label": "Terrenos", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "project_management_coste", "item_label": "Project Management", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "acometidas", "item_label": "Acometidas", "is_percent": False},
            {"group_name": "Coste comprador", "item_key": "costes_construccion", "item_label": "Costes de construcci√≥n", "is_percent": False},
        ],
        "R2B + PM": [
            # Same as R2B for now
        ],
        "R2B + PM + Venta certs": [
            # Same as R2B for now
        ],
        "Promoci√≥n": [
            # Different structure for Promoci√≥n
        ]
    }
    
    structure = template_structures.get(template_key, template_structures["R2B"])
    
    # For now, just return success - the structure will be created when the user sets values
    # The RPC list_property_numbers should handle this, but if it doesn't, we'll need to create an RPC
    return {"property_id": property_id, "template_key": template_key, "status": "structure_initialized"}

def clear_numbers(property_id: str) -> Dict:
    """Clear/reset all number values for a property when starting fresh with a new template.
    
    This function attempts to clear all number values, but if the RPC doesn't exist,
    it will return success anyway (template selection is the priority).
    """
    # Try RPC first (if it exists) to avoid 404 errors from direct table access
    try:
        sb.postgrest.schema = "public"
        sb.rpc("clear_property_numbers", {"p_id": property_id}).execute()
        return {"property_id": property_id, "status": "cleared"}
    except Exception:
        # If RPC doesn't exist, return success anyway (template selection is the priority)
        # The values will be cleared when the user starts entering new values in the Excel
        # This avoids 404 errors in the logs while still allowing template selection to proceed
        return {"property_id": property_id, "status": "attempted", "note": "RPC not available, values will be cleared when user enters new values"}

def find_item_by_value(property_id: str, search_value: Optional[Union[str, float]] = None, search_label: Optional[str] = None) -> Optional[Dict]:
    """Find an item in the numbers framework by value or label.
    
    Args:
        property_id: Property UUID
        search_value: Value to search for (e.g., 10.0 for "10%" or 100000)
        search_label: Label text to search for (e.g., "IVA", "Precio de venta")
    
    Returns:
        Dict with item_key, item_label, amount, etc. or None if not found
    """
    items = get_numbers(property_id)
    
    # Normalize search_label for fuzzy matching
    if search_label:
        search_label_lower = search_label.lower().strip()
        # Remove common words and normalize, but keep "iva" for matching
        search_label_clean = search_label_lower.replace("(%)", "").replace("(", "").replace(")", "").strip()
    
    # Parse search_value if it's a string
    parsed_value = None
    if search_value is not None:
        if isinstance(search_value, (int, float)):
            parsed_value = float(search_value)
        elif isinstance(search_value, str):
            try:
                val_str = str(search_value).replace("%", "").replace(",", ".").strip()
                parsed_value = float(val_str)
            except:
                pass
    
    best_match = None
    best_score = 0
    
    for item in items:
        item_label = item.get("item_label", "").lower()
        item_key = item.get("item_key", "").lower()
        item_amount = item.get("amount")
        
        score = 0
        
        # Match by label (fuzzy)
        if search_label:
            # Exact match in label
            if search_label_clean in item_label or item_label in search_label_clean:
                score += 10
            # Partial match in label
            elif any(word in item_label for word in search_label_clean.split() if len(word) > 2):
                score += 5
            # Match in item_key
            if search_label_clean in item_key:
                score += 8
        
        # Match by value
        if parsed_value is not None and item_amount is not None:
            # Allow small floating point differences
            if abs(float(item_amount) - parsed_value) < 0.01:
                score += 10
            # Allow percentage matching (e.g., 10% = 10.0)
            elif abs(float(item_amount) - parsed_value) < 1.0:
                score += 5
        
        # If both label and value match, return immediately (perfect match)
        if score >= 20:
            return item
        
        # Track best match
        if score > best_score:
            best_score = score
            best_match = item
    
    # Return best match if score is high enough
    if best_score >= 10:
        return best_match
    
    # Fallback: if only label or only value matches, return it
    if search_label and best_score >= 5:
        return best_match
    if parsed_value is not None and best_score >= 5:
        return best_match
    
    return None


# ==================== NEW: Numbers Table Framework (Excel Replica) ====================

def _rgb_to_hex(rgb_obj) -> str | None:
    """Convert openpyxl RGB object to hex string for JSON serialization."""
    if not rgb_obj:
        return None
    try:
        # openpyxl RGB objects have a .rgb attribute
        if hasattr(rgb_obj, 'rgb'):
            rgb_hex = rgb_obj.rgb
            if isinstance(rgb_hex, str):
                # Format: "FFRRGGBB" -> "#RRGGBB"
                if rgb_hex.startswith('FF') and len(rgb_hex) == 8:
                    return '#' + rgb_hex[2:]
                elif not rgb_hex.startswith('#'):
                    return '#' + rgb_hex if len(rgb_hex) == 6 else rgb_hex
                return rgb_hex
        # Fallback
        return str(rgb_obj) if rgb_obj else None
    except Exception:
        return None


def import_excel_from_file(file_bytes: bytes, property_id: str, template_key: str) -> Dict:
    """Import Excel template structure from uploaded file bytes using openpyxl.
    This is an alternative to Graph API that doesn't require authentication.
    
    Args:
        file_bytes: Excel file bytes
        property_id: Property UUID
        template_key: Template identifier (e.g., "R2B")
    
    Returns:
        Dict with structure_json and imported cell count
    """
    from openpyxl import load_workbook
    from io import BytesIO
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Starting Excel import from file: template_key={template_key}, property_id={property_id}")
        
        # Load workbook from bytes
        # CRITICAL: data_only=False to read formulas, not just values
        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        ws = wb.active  # Use first worksheet
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        logger.info(f"Excel dimensions: {max_row} rows x {max_col} columns")
        
        # Build structure
        structure = {
            "rows": max_row,
            "columns": max_col,
            "cells": []
        }
        
        # Extract header row (first row)
        header_row = []
        for col in range(1, min(max_col + 1, 27)):  # A-Z
            cell = ws.cell(row=1, column=col)
            header_row.append(str(cell.value) if cell.value else "")
        structure["header_row"] = header_row
        
        # Extract header column (first column)
        header_col = []
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=1)
            header_col.append(str(cell.value) if cell.value else "")
        structure["header_col"] = header_col
        
        # Extract ALL cells (including empty ones) to preserve exact Excel structure
        cell_values = {}
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # Convert column number to letter (A, B, ..., Z, AA, AB, ...)
                col_letter = ""
                col_num = col
                while col_num > 0:
                    col_num -= 1
                    col_letter = chr(65 + (col_num % 26)) + col_letter
                    col_num //= 26
                
                cell_address = f"{col_letter}{row}"
                
                # Get value (even if empty)
                value = cell.value
                formula = None
                if cell.data_type == 'f':
                    formula = cell.value  # Formula string
                    # Try to get calculated value
                    try:
                        value = cell.value if hasattr(cell, 'value') else None
                    except:
                        value = None
                
                # Get format (convert RGB objects to strings for JSON serialization)
                cell_format = {}
                if cell.fill and cell.fill.start_color:
                    bg_color = _rgb_to_hex(cell.fill.start_color.rgb)
                    if bg_color:
                        cell_format["bg_color"] = bg_color
                
                if cell.font:
                    if cell.font.color:
                        font_color = _rgb_to_hex(cell.font.color.rgb)
                        if font_color:
                            cell_format["font_color"] = font_color
                    cell_format["bold"] = bool(cell.font.bold) if cell.font else False
                else:
                    cell_format["bold"] = False
                
                # Get borders if present
                if cell.border:
                    borders = {}
                    if cell.border.left:
                        borders["left"] = str(cell.border.left.style) if cell.border.left.style else None
                    if cell.border.right:
                        borders["right"] = str(cell.border.right.style) if cell.border.right.style else None
                    if cell.border.top:
                        borders["top"] = str(cell.border.top.style) if cell.border.top.style else None
                    if cell.border.bottom:
                        borders["bottom"] = str(cell.border.bottom.style) if cell.border.bottom.style else None
                    if borders:
                        cell_format["borders"] = borders
                
                # Get alignment if present
                if cell.alignment:
                    alignment = {}
                    if cell.alignment.horizontal:
                        alignment["horizontal"] = str(cell.alignment.horizontal)
                    if cell.alignment.vertical:
                        alignment["vertical"] = str(cell.alignment.vertical)
                    if alignment:
                        cell_format["alignment"] = alignment
                
                # Get row and column labels
                row_label = None
                col_label = None
                if row > 1:  # Skip header row
                    row_label_cell = ws.cell(row=row, column=1)
                    row_label = str(row_label_cell.value) if row_label_cell.value else None
                if col > 1:  # Skip header column
                    col_label = header_row[col - 1] if col - 1 < len(header_row) else None
                
                # Identify if cell is a user input (yellow background)
                is_user_input = False
                if cell_format.get("bg_color"):
                    # Check if background is yellow/yellowish (FFFF00, FFFFE0, etc.)
                    bg = cell_format["bg_color"].upper()
                    if bg.startswith("FFFF") or "YELLOW" in bg:
                        is_user_input = True
                
                # Store ALL cells in structure (including empty ones)
                # CRITICAL: If cell has a formula (or will have one injected), DON'T store its calculated value
                # This prevents showing 0 in cells that should be empty until calculated
                cell_value_to_store = ""
                
                # Check if this cell will have a formula injected (R2B template)
                will_have_formula = (template_key == "R2B" and cell_address in R2B_FORMULAS)
                
                if (formula and isinstance(formula, str) and formula.startswith("=")) or will_have_formula:
                    # Has formula (or will have one) - don't store value, only formula
                    cell_value_to_store = ""
                else:
                    # No formula - store the actual value
                    cell_value_to_store = str(value) if value is not None else ""
                
                structure["cells"].append({
                    "address": cell_address,
                    "row": row,
                    "col": col,
                    "col_letter": col_letter,
                    "value": cell_value_to_store,
                    "formula": formula if formula and isinstance(formula, str) and formula.startswith("=") else None,
                    "format": cell_format,
                    "row_label": row_label,
                    "col_label": col_label,
                    "is_user_input": is_user_input  # Mark yellow cells as user inputs
                })
                
                # Store for values table (only if has value or formula)
                if value is not None or formula:
                    cell_values[cell_address] = {
                        "value": str(value) if value is not None else "",
                        "row_label": row_label,
                        "col_label": col_label,
                        "format": cell_format
                    }
        
        # Save structure to numbers_templates
        # Use the already imported sb from supabase_client
        sb.postgrest.schema = "public"
        
        template_data = {
            "template_key": template_key,
            "property_id": property_id,
            "structure_json": structure
        }
        
        # Upsert template
        existing = sb.table("numbers_templates").select("id").eq("template_key", template_key).eq("property_id", property_id).execute()
        if existing.data:
            sb.table("numbers_templates").update(template_data).eq("id", existing.data[0]["id"]).execute()
        else:
            sb.table("numbers_templates").insert(template_data).execute()
        
        # Save initial values to numbers_table_values
        # CR√çTICO: NO sobrescribir valores que el usuario ya a√±adi√≥ via chat
        # CR√çTICO: NO guardar valores para celdas que tienen f√≥rmulas en R2B_FORMULAS
        # Solo guardar valores que vienen del Excel y que NO existen en la DB y NO tienen f√≥rmulas
        existing_values = get_numbers_table_values(property_id, template_key)
        saved_count = 0
        skipped_count = 0
        formula_cells_skipped = 0
        
        for cell_addr, cell_data in cell_values.items():
            try:
                normalized_addr = str(cell_addr).upper().strip()
                
                # Skip cells that have formulas in R2B_FORMULAS (they should remain empty in DB)
                if template_key == "R2B" and normalized_addr in R2B_FORMULAS:
                    logger.info(f"[import_excel] Skipping {normalized_addr} - has formula in R2B_FORMULAS")
                    formula_cells_skipped += 1
                    continue
                
                # Check if user already added a value for this cell via chat
                if normalized_addr in existing_values:
                    existing_val = existing_values[normalized_addr]
                    existing_val_str = existing_val.get("value", "") if isinstance(existing_val, dict) else str(existing_val)
                    # Skip if there's already a user-added value (non-empty)
                    if existing_val_str and existing_val_str != cell_data["value"]:
                        logger.info(f"[import_excel] Preserving user value for {normalized_addr}: '{existing_val_str}' (Excel has '{cell_data['value']}')")
                        skipped_count += 1
                        continue
                
                sb.rpc("set_numbers_table_cell", {
                    "p_property_id": property_id,
                    "p_template_key": template_key,
                    "p_cell_address": cell_addr,
                    "p_value": cell_data["value"],
                    "p_row_label": cell_data.get("row_label"),
                    "p_col_label": cell_data.get("col_label"),
                    "p_format_json": cell_data.get("format", {})
                }).execute()
                saved_count += 1
            except Exception as e:
                logger.warning(f"Failed to save cell {cell_addr}: {e}")
        
        logger.info(f"‚úÖ Imported {saved_count} cells from Excel file, preserved {skipped_count} user-added values, skipped {formula_cells_skipped} formula cells")
        
        # üî• INJECT R2B FORMULAS AUTOMATICALLY INTO STRUCTURE
        if template_key == "R2B":
            logger.info(f"[R2B] üî• Injecting {len(R2B_FORMULAS)} formulas into structure...")
            injected_count = 0
            
            # Update structure with formulas
            for cell in structure.get("cells", []):
                cell_address = cell.get("address")
                if cell_address in R2B_FORMULAS:
                    cell["formula"] = R2B_FORMULAS[cell_address]
                    cell["value"] = ""  # CRITICAL: Clear any stored value for formula cells
                    injected_count += 1
                    logger.debug(f"[R2B] Injected formula: {cell_address} = {R2B_FORMULAS[cell_address]}")
            
            # Re-save structure with formulas
            existing = sb.table("numbers_templates").select("id").eq("template_key", template_key).eq("property_id", property_id).execute()
            if existing.data:
                sb.table("numbers_templates").update({
                    "structure_json": structure
                }).eq("id", existing.data[0]["id"]).execute()
                logger.info(f"[R2B] ‚úÖ Successfully injected {injected_count}/{len(R2B_FORMULAS)} formulas into structure")
            else:
                logger.warning(f"[R2B] ‚ö†Ô∏è Could not find template to update with formulas")
        
        return {
            "ok": True,
            "template_key": template_key,
            "property_id": property_id,
            "structure": structure,
            "cells_imported": saved_count
        }
        
    except Exception as e:
        logger.error(f"Error importing Excel from file: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


def import_excel_template(property_id: str, template_key: str, excel_file_id: str, access_token: str) -> Dict:
    """Import Excel template structure from Microsoft Graph API and save to numbers_templates table.
    This is a fallback method when file upload is not available.
    """
    import requests
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Starting Excel import via Graph API: template_key={template_key}, property_id={property_id}")
        # This function would use Graph API - but we prefer file upload
        return {"ok": False, "error": "Graph API import not available. Please upload the Excel file directly."}
    except Exception as e:
        logger.error(f"Error importing Excel template: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


def get_numbers_table_structure(property_id: str, template_key: str) -> Dict:
    """Get the structure JSON for a Numbers template."""
    try:
        sb.postgrest.schema = "public"
        result = sb.rpc("get_numbers_template_structure", {
            "p_template_key": template_key,
            "p_property_id": property_id
        }).execute()
        if result.data:
            if isinstance(result.data, dict):
                if "rows" in result.data or "cells" in result.data:
                    return result.data
            return result.data if result.data else {}
        return {}
    except Exception as e:
        import logging
        logging.error(f"Error getting template structure: {e}")
        return {}


def get_numbers_table_values(property_id: str, template_key: str) -> Dict:
    """Get all cell values for a property's Numbers table."""
    import logging
    logger = logging.getLogger(__name__)
    try:
        sb.postgrest.schema = "public"
        logger.info(f"[get_numbers_table_values] Fetching values for property_id={property_id}, template_key={template_key}")
        result = sb.rpc("get_numbers_table_values", {
            "p_property_id": property_id,
            "p_template_key": template_key
        }).execute()
        values = result.data if result.data else {}
        logger.info(f"[get_numbers_table_values] Retrieved {len(values)} values from DB")
        if values:
            logger.info(f"[get_numbers_table_values] Sample values: {dict(list(values.items())[:5])}")
        return values
    except Exception as e:
        logger.error(f"[get_numbers_table_values] Error getting table values: {e}", exc_info=True)
        return {}


def set_numbers_table_cell(property_id: str, template_key: str, cell_address: str, value: str, row_label: Optional[str] = None, col_label: Optional[str] = None, format_json: Optional[Dict] = None, auto_calculate: bool = True) -> Dict:
    """Set a cell value in the Numbers table with automatic formula calculation.
    
    üî• C√ÅLCULO AUTOM√ÅTICO EN CASCADA:
    - Cuando auto_calculate=True (default), despu√©s de actualizar la celda,
      se recalculan autom√°ticamente todas las f√≥rmulas dependientes en cascada.
    - Ejemplo: Si actualizas B5 y C5, D5 (=B5*C5/100) se calcular√° autom√°ticamente,
      y luego E5 (=B5+D5) tambi√©n se calcular√° en cascada.
    
    Returns validated result with auto_calculated cells if any.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        sb.postgrest.schema = "public"
        # Normalize cell address to uppercase for consistent storage
        normalized_cell_address = str(cell_address).upper().strip()
        logger.info(f"[set_numbers_table_cell] Setting {normalized_cell_address} = '{value}' for property_id={property_id}, template_key={template_key}, auto_calculate={auto_calculate}")
        
        # Ensure property_id is a valid UUID string
        if not property_id:
            raise ValueError("property_id is required")
        
        # 1. Save the user's value to DB
        result = sb.rpc("set_numbers_table_cell", {
            "p_property_id": property_id,
            "p_template_key": template_key,
            "p_cell_address": normalized_cell_address,
            "p_value": str(value) if value is not None else "",
            "p_row_label": row_label,
            "p_col_label": col_label,
            "p_format_json": format_json or {}
        }).execute()
        
        logger.info(f"[set_numbers_table_cell] RPC result: {result.data}")
        
        # 2. Validate that the value was saved correctly
        if result.data and result.data.get("ok"):
            # Verify by reading back the value (use normalized address)
            values = get_numbers_table_values(property_id, template_key)
            saved_value = values.get(normalized_cell_address, {})
            if isinstance(saved_value, dict):
                saved_value_str = saved_value.get("value", "")
            else:
                saved_value_str = str(saved_value)
            
            expected_value_str = str(value) if value is not None else ""
            if saved_value_str == expected_value_str:
                logger.info(f"‚úÖ Validated: {normalized_cell_address} = {value} saved correctly")
                
                response = {"ok": True, "cell_address": normalized_cell_address, "value": value, "validated": True}
                
                # 3. Auto-calculate dependent formulas if enabled
                if auto_calculate:
                    logger.info(f"[set_numbers_table_cell] üî• Starting auto-calculation for {normalized_cell_address}")
                    
                    try:
                        # Get template structure
                        structure = get_numbers_table_structure(property_id, template_key)
                        
                        # Get current values from DB
                        current_values = values  # We already fetched this above
                        
                        # Calculate affected cells
                        calculated = auto_calculate_on_update(
                            property_id=property_id,
                            template_key=template_key,
                            updated_cell=normalized_cell_address,
                            new_value=value,
                            structure=structure,
                            current_values=current_values
                        )
                        
                        # 4. Save calculated values to DB so frontend can display them
                        # Excel export will still use formulas from structure (not these stored values)
                        if calculated:
                            logger.info(f"[set_numbers_table_cell] ‚úÖ Auto-calculated {len(calculated)} cells, saving to DB for frontend display")
                            
                            # Save each calculated value to DB
                            for calc_cell_address, calc_value in calculated.items():
                                try:
                                    # Find cell info for this calculated cell
                                    calc_cell_info = None
                                    for cell in structure.get("cells", []):
                                        if cell.get("address") == calc_cell_address:
                                            calc_cell_info = cell
                                            break
                                    
                                    # Save via RPC
                                    sb.rpc("set_numbers_table_cell", {
                                        "p_property_id": property_id,
                                        "p_template_key": template_key,
                                        "p_cell_address": calc_cell_address,
                                        "p_value": str(calc_value),
                                        "p_row_label": calc_cell_info.get("row_label") if calc_cell_info else None,
                                        "p_col_label": calc_cell_info.get("col_label") if calc_cell_info else None,
                                        "p_format_json": calc_cell_info.get("format", {}) if calc_cell_info else {}
                                    }).execute()
                                    logger.debug(f"[set_numbers_table_cell] Saved calculated value: {calc_cell_address} = {calc_value}")
                                except Exception as e:
                                    logger.warning(f"[set_numbers_table_cell] Error saving calculated value for {calc_cell_address}: {e}")
                            
                            response["auto_calculated"] = calculated
                            response["auto_calculated_count"] = len(calculated)
                            logger.info(f"‚úÖ Auto-calculated and saved {len(calculated)} cells: {list(calculated.keys())}")
                        else:
                            logger.info(f"[set_numbers_table_cell] No formulas to calculate for {normalized_cell_address}")
                    
                    except Exception as calc_error:
                        logger.error(f"[set_numbers_table_cell] ‚ö†Ô∏è Error during auto-calculation: {calc_error}", exc_info=True)
                        # Don't fail the whole operation if auto-calc fails
                        response["auto_calculate_error"] = str(calc_error)
                
                # Log-only verification
                try:
                    verify_numbers_update(property_id, template_key, normalized_cell_address, str(value), response.get("auto_calculated"))
                except Exception:
                    pass
                # Metrics event
                # Log to Logfire
                import logfire
                logfire.info("numbers_cell_updated", property_id=property_id, cell=normalized_cell_address, value=str(value)[:50])
                return response
            else:
                logger.warning(f"‚ö†Ô∏è Validation failed: expected {expected_value_str}, got {saved_value_str}")
                return {"ok": False, "cell_address": normalized_cell_address, "error": f"Validation failed: expected {expected_value_str}, got {saved_value_str}"}
        
        return result.data if result.data else {"ok": True, "cell_address": normalized_cell_address, "value": value, "validated": False}
    except Exception as e:
        logger.error(f"Error setting cell value {cell_address}: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


def clear_numbers_table_cell(property_id: str, template_key: str, cell_address: str) -> Dict:
    """Clear/delete a cell value in the Numbers table. Returns validated result."""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        sb.postgrest.schema = "public"
        # Delete the cell value by setting it to empty string (or we can delete the row)
        # Using set_numbers_table_cell with empty value to clear it
        result = sb.rpc("set_numbers_table_cell", {
            "p_property_id": property_id,
            "p_template_key": template_key,
            "p_cell_address": cell_address,
            "p_value": "",  # Empty string to clear the value
            "p_row_label": None,
            "p_col_label": None,
            "p_format_json": {}
        }).execute()
        
        # Validate that the value was deleted correctly
        if result.data and result.data.get("ok"):
            # Verify by reading back the value - it should be empty or not exist
            values = get_numbers_table_values(property_id, template_key)
            saved_value = values.get(cell_address, {})
            if isinstance(saved_value, dict):
                saved_value_str = saved_value.get("value", "")
            else:
                saved_value_str = str(saved_value) if saved_value else ""
            
            # Value should be empty string or not exist
            if saved_value_str == "" or saved_value_str is None:
                logger.info(f"‚úÖ Validated: {cell_address} cleared successfully")
                return {"ok": True, "cell_address": cell_address, "cleared": True, "validated": True}
            else:
                logger.warning(f"‚ö†Ô∏è Validation failed: expected empty, got {saved_value_str}")
                return {"ok": False, "cell_address": cell_address, "error": f"Validation failed: expected empty, got {saved_value_str}"}
        
        return result.data if result.data else {"ok": True, "cell_address": cell_address, "cleared": True, "validated": False}
    except Exception as e:
        logger.error(f"Error clearing cell value {cell_address}: {e}")
        return {"ok": False, "error": str(e)}


def delete_numbers_template(property_id: str, template_key: str = "R2B") -> Dict:
    """Delete a Numbers template completely (structure and all values).
    
    This is useful when you need to re-import the template cleanly.
    
    Args:
        property_id: Property UUID
        template_key: Template identifier (default "R2B")
    
    Returns:
        Dict with ok status and counts of deleted records
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        sb.postgrest.schema = "public"
        logger.info(f"[delete_numbers_template] Deleting template {template_key} for property {property_id}")
        
        # Delete from numbers_templates
        result1 = sb.table("numbers_templates").delete().eq("property_id", property_id).eq("template_key", template_key).execute()
        deleted_templates = len(result1.data) if result1.data else 0
        logger.info(f"[delete_numbers_template] Deleted {deleted_templates} template(s)")
        
        # Delete from numbers_table_values
        result2 = sb.table("numbers_table_values").delete().eq("property_id", property_id).eq("template_key", template_key).execute()
        deleted_values = len(result2.data) if result2.data else 0
        logger.info(f"[delete_numbers_template] Deleted {deleted_values} value(s)")
        
        return {
            "ok": True,
            "deleted_templates": deleted_templates,
            "deleted_values": deleted_values,
            "message": f"Template {template_key} deleted successfully"
        }
    except Exception as e:
        logger.error(f"[delete_numbers_template] Error: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}
