from __future__ import annotations
import io
import json
from typing import Dict, Tuple, Any, List

from .numbers_tools import get_numbers
from .supabase_client import sb
from .supabase_client import BUCKET


def _to_map(rows: list[dict]) -> Dict[str, float]:
    """Convert numbers rows into a simple {item_key: amount} map, skipping None.
    Percent fields are already stored as 0.xx according to product spec.
    """
    out: Dict[str, float] = {}
    for r in rows or []:
        k = r.get("item_key")
        amt = r.get("amount")
        if k is None:
            continue
        try:
            out[k] = float(amt) if amt is not None else None
        except Exception:
            out[k] = None
    return out


def _safe_div(a: float | None, b: float | None) -> float | None:
    if a is None or b is None or b == 0:
        return None
    return a / b


def compute_derived_from_inputs(inputs: Dict[str, float]) -> Dict[str, float | None]:
    """Compute derived metrics using provided inputs.
    Keys expected (optional if missing):
    - precio_venta, impuestos_pct, project_mgmt_fees, terrenos_coste,
      project_management_coste, acometidas, costes_construccion,
      total_pagado, terreno_urbano, terreno_rustico, superficie_m2 (optional)
    """
    precio_venta = inputs.get("precio_venta")
    impuestos_pct = inputs.get("impuestos_pct")
    project_mgmt_fees = inputs.get("project_mgmt_fees")
    terrenos_coste = inputs.get("terrenos_coste")
    project_management_coste = inputs.get("project_management_coste")
    acometidas = inputs.get("acometidas")
    costes_construccion = inputs.get("costes_construccion")
    total_pagado = inputs.get("total_pagado")
    terreno_urbano = inputs.get("terreno_urbano")
    terreno_rustico = inputs.get("terreno_rustico")
    superficie_m2 = inputs.get("superficie_m2")

    impuestos_total = None
    if impuestos_pct is not None and precio_venta is not None:
        impuestos_total = impuestos_pct * precio_venta

    def _sum(values: list[float | None]) -> float | None:
        acc = 0.0
        has = False
        for v in values:
            if v is not None:
                has = True
                acc += float(v)
        return acc if has else None

    costes_totales = _sum([
        project_mgmt_fees,
        terrenos_coste,
        project_management_coste,
        acometidas,
        costes_construccion,
    ])

    gross_margin = None
    if precio_venta is not None and costes_totales is not None:
        gross_margin = precio_venta - costes_totales

    net_profit = None
    if precio_venta is not None and costes_totales is not None and impuestos_total is not None:
        net_profit = precio_venta - costes_totales - impuestos_total

    roi_pct = _safe_div(net_profit, total_pagado)

    urbano_ratio = None
    if terreno_urbano is not None and terreno_rustico is not None:
        denom = terreno_urbano + terreno_rustico
        urbano_ratio = _safe_div(terreno_urbano, denom)

    price_per_m2 = _safe_div(precio_venta, superficie_m2)

    return {
        "impuestos_total": impuestos_total,
        "costes_totales": costes_totales,
        "gross_margin": gross_margin,
        "net_profit": net_profit,
        "roi_pct": roi_pct,
        "urbano_ratio": urbano_ratio,
        "price_per_m2": price_per_m2,
    }


def validate_anomalies(inputs: Dict[str, float], outputs: Dict[str, float | None]) -> list[str]:
    warnings: list[str] = []
    impuestos_pct = inputs.get("impuestos_pct")
    precio_venta = inputs.get("precio_venta")
    total_pagado = inputs.get("total_pagado")
    net_profit = outputs.get("net_profit")

    # impuestos_pct range
    if impuestos_pct is not None and not (0 <= impuestos_pct <= 0.25):
        warnings.append("impuestos_pct fuera de rango [0,0.25]")
    # non-negative checks (selected inputs)
    for k in [
        "precio_venta", "project_mgmt_fees", "terrenos_coste", "project_management_coste",
        "acometidas", "costes_construccion", "total_pagado", "terreno_urbano", "terreno_rustico"
    ]:
        v = inputs.get(k)
        if v is not None and v < 0:
            warnings.append(f"{k} es negativo")
    # total_pagado vs precio_venta
    if precio_venta is not None and total_pagado is not None and total_pagado > precio_venta:
        warnings.append("total_pagado > precio_venta")
    # net_profit negative
    if net_profit is not None and net_profit < 0:
        warnings.append("net_profit negativo")

    return warnings


def compute_and_log(property_id: str, triggered_by: str = "agent", trigger_type: str = "manual") -> Dict[str, Any]:
    """Compute derived metrics for a property, persist best-effort to calc_outputs and calc_log.
    Returns {inputs, outputs, anomalies}.
    """
    rows = get_numbers(property_id)
    inputs = _to_map(rows)
    outputs = compute_derived_from_inputs(inputs)
    anomalies = validate_anomalies(inputs, outputs)

    # Best-effort persistence (tables may not exist yet)
    try:
        sb.table("calc_outputs").upsert({
            "property_id": property_id,
            "outputs": outputs,
            "anomalies": anomalies,
        }, on_conflict="property_id").execute()
    except Exception:
        pass
    try:
        sb.table("calc_log").insert({
            "property_id": property_id,
            "inputs": inputs,
            "outputs": outputs,
            "anomalies": anomalies,
            "triggered_by": triggered_by,
            "trigger_type": trigger_type,
        }).execute()
    except Exception:
        pass

    return {"inputs": inputs, "outputs": outputs, "anomalies": anomalies}


def generate_numbers_excel(property_id: str) -> bytes:
    """Create an Excel workbook with Inputs, Derived, and Anomalies sheets. Returns bytes."""
    import pandas as pd
    # Compute fresh values for the export
    result = compute_and_log(property_id, triggered_by="agent", trigger_type="export")
    inputs = result["inputs"]
    outputs = result["outputs"]
    anomalies = result["anomalies"]
    # Try to include last sensitivity or what-if snapshots if present
    scenarios_df = None
    sens_df = None
    try:
        snaps = sb.table("scenario_snapshots").select("name,deltas,outputs,created_at").eq("property_id", property_id).order("created_at", desc=True).limit(50).execute().data
        if snaps:
            rows = []
            for s in snaps:
                rows.append({
                    "name": s.get("name"),
                    "deltas": s.get("deltas"),
                    "outputs": s.get("outputs"),
                    "created_at": s.get("created_at"),
                })
            scenarios_df = pd.json_normalize(rows)
            # Extract last sensitivity grid if any
            for s in snaps:
                if (s.get("name") or "").lower() == "sensitivity" and s.get("outputs", {}).get("grid"):
                    z = s["outputs"]["grid"]
                    sens_df = pd.DataFrame(z)
                    break
    except Exception:
        pass

    # Build dataframes
    df_inputs = pd.DataFrame([{"item_key": k, "amount": inputs.get(k)} for k in sorted(inputs.keys())])
    df_outputs = pd.DataFrame([{"metric": k, "value": outputs.get(k)} for k in sorted(outputs.keys())])
    df_anom = pd.DataFrame({"anomaly": anomalies}) if anomalies else pd.DataFrame(columns=["anomaly"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_inputs.to_excel(writer, index=False, sheet_name="Inputs")
        df_outputs.to_excel(writer, index=False, sheet_name="Derived")
        df_anom.to_excel(writer, index=False, sheet_name="Anomalies")
        if scenarios_df is not None:
            scenarios_df.to_excel(writer, index=False, sheet_name="Scenarios")
        if sens_df is not None:
            sens_df.to_excel(writer, index=False, sheet_name="SensitivityGrid")
    return buf.getvalue()


def generate_numbers_table_excel(property_id: str, template_key: str = "R2B") -> bytes:
    """Create an Excel workbook from the Numbers table structure and values.
    Recreates the exact Excel structure with headers, labels, format, and values.
    Returns bytes."""
    from .numbers_tools import get_numbers_table_structure, get_numbers_table_values
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.borders import BORDER_THIN, BORDER_MEDIUM, BORDER_THICK
    import io
    
    try:
        # Get structure and values
        structure = get_numbers_table_structure(property_id, template_key)
        values = get_numbers_table_values(property_id, template_key)
        
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[generate_numbers_table_excel] property_id={property_id}, template_key={template_key}")
        logger.info(f"[generate_numbers_table_excel] structure keys: {list(structure.keys()) if structure else 'None'}")
        logger.info(f"[generate_numbers_table_excel] structure cells count: {len(structure.get('cells', [])) if structure else 0}")
        logger.info(f"[generate_numbers_table_excel] values from DB: {values}")
        logger.info(f"[generate_numbers_table_excel] values count: {len(values)}")
        logger.info(f"[generate_numbers_table_excel] Sample values: {dict(list(values.items())[:5]) if values else 'No values'}")
        
        if not structure or not structure.get("cells"):
            # DO NOT fallback to old method - raise error instead
            error_msg = f"No se encontró la estructura de la plantilla {template_key} para la propiedad {property_id}. Por favor, importa primero la plantilla Excel."
            logger.error(f"[generate_numbers_table_excel] {error_msg}")
            raise ValueError(error_msg)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Build cell map from values (values from DB - these are the updated values added via chat)
        # CRITICAL: Normalize cell addresses to uppercase for consistent matching
        cell_map = {}
        for cell_addr, cell_data in values.items():
            # Normalize cell address to uppercase (e.g., "b5" -> "B5")
            normalized_addr = str(cell_addr).upper().strip()
            
            if isinstance(cell_data, dict):
                cell_value = cell_data.get("value", "")
                # Store value even if empty string (to distinguish from None)
                cell_map[normalized_addr] = {
                    "value": cell_value if cell_value is not None else "",
                    "format": cell_data.get("format", {})
                }
                logger.info(f"[generate_numbers_table_excel] Loaded value from DB: {normalized_addr} = '{cell_value}' (original: {cell_addr})")
            else:
                # Legacy format: just value
                cell_map[normalized_addr] = {"value": str(cell_data), "format": {}}
                logger.info(f"[generate_numbers_table_excel] Loaded value from DB (legacy): {normalized_addr} = '{cell_data}' (original: {cell_addr})")
        
        logger.info(f"[generate_numbers_table_excel] Total values loaded from DB: {len(cell_map)}")
        logger.info(f"[generate_numbers_table_excel] Cell addresses in DB: {sorted(list(cell_map.keys()))[:10]}")
        
        # Create a map of all cells by address for quick lookup
        structure_cell_map = {}
        for cell_info in structure.get("cells", []):
            cell_addr = cell_info.get("address")
            structure_cell_map[cell_addr] = cell_info
        
        # Write ALL cells from structure (preserve exact structure including empty cells)
        cells_with_db_values = 0
        cells_with_structure_values = 0
        cells_with_formulas = 0
        cells_empty = 0
        
        for cell_info in structure.get("cells", []):
            cell_addr = cell_info.get("address")
            row = cell_info.get("row", 1)
            col = cell_info.get("col", 1)
            col_letter = cell_info.get("col_letter") or get_column_letter(col)
            
            cell = ws[f"{col_letter}{row}"]
            
            # Normalize cell address to uppercase for consistent matching
            normalized_cell_addr = str(cell_addr).upper().strip()
            
            # Check if there's an updated value in cell_map (from DB)
            updated_value_data = cell_map.get(normalized_cell_addr)
            
            # DEBUG: Log if we're looking for a specific cell
            if normalized_cell_addr in ["B5", "C10", "D8"]:  # Common test cells
                logger.info(f"[generate_numbers_table_excel] 🔍 Checking cell {normalized_cell_addr}: DB data = {updated_value_data}, structure value = {cell_info.get('value')}")
            
            # Priority: 1) Updated value from DB (if non-empty), 2) Formula from structure, 3) Original value from structure
            value_written = False
            
            # Check DB value first, BUT skip if this cell has a formula in the structure
            # (we want to use formulas in Excel, not stored calculated values)
            has_formula = bool(cell_info.get("formula"))
            
            if updated_value_data is not None and not has_formula:
                # We have a value from DB and this cell doesn't have a formula
                if isinstance(updated_value_data, dict):
                    value = updated_value_data.get("value", "")
                else:
                    value = str(updated_value_data)
                
                # Only write non-empty values from DB
                if value and isinstance(value, str) and value.strip():
                    logger.info(f"[generate_numbers_table_excel] ✅ DB value found for {normalized_cell_addr}: '{value}'")
                    try:
                        if not value.startswith("="):
                            # Try to convert to number
                            num_val = float(value.replace(",", "."))
                            cell.value = num_val
                            logger.info(f"[generate_numbers_table_excel] ✅ Set {normalized_cell_addr} to number: {num_val}")
                            cells_with_db_values += 1
                            value_written = True
                        else:
                            cell.value = value
                            logger.info(f"[generate_numbers_table_excel] ✅ Set {normalized_cell_addr} to formula: '{value}'")
                            cells_with_db_values += 1
                            value_written = True
                    except (ValueError, TypeError) as e:
                        cell.value = str(value)
                        logger.warning(f"[generate_numbers_table_excel] ⚠️ Error converting value for {normalized_cell_addr}: {e}, set to string: '{value}'")
                        cells_with_db_values += 1
                        value_written = True
                else:
                    logger.info(f"[generate_numbers_table_excel] ⏭️ Skipping empty DB value for {normalized_cell_addr}, will use formula or structure")
            elif updated_value_data is not None and has_formula:
                logger.debug(f"[generate_numbers_table_excel] ⏭️ Ignoring DB value for {normalized_cell_addr} (cell has formula, will use formula instead)")
            
            # If no value from DB, check formula
            if not value_written:
                cell_formula = cell_info.get("formula")
                # DEBUG: Log for known formula cells
                if normalized_cell_addr in ["B10", "D5", "D6", "D7", "D8", "E5", "B12", "B13", "B14", "B15"]:
                    logger.info(f"[generate_numbers_table_excel] 🔍 Formula cell {normalized_cell_addr}: formula in structure = '{cell_formula}', value in structure = '{cell_info.get('value')}'")
                
                if cell_formula:
                    # Preserve formula from original Excel
                    cell.value = cell_formula
                    logger.info(f"[generate_numbers_table_excel] ✅ Using formula for {normalized_cell_addr}: {cell_formula}")
                    cells_with_formulas += 1
                    value_written = True
            
            # If no value and no formula, check structure value ONLY if cell doesn't have a formula defined
            if not value_written:
                # CRITICAL: Check if this cell is a formula cell in R2B_FORMULAS
                # If yes, DON'T use structure value (it's likely a calculated 0)
                from tools.numbers_tools import R2B_FORMULAS
                is_formula_cell = template_key == "R2B" and normalized_cell_addr in R2B_FORMULAS
                
                if not is_formula_cell:
                    # Safe to use original value from structure (not a formula cell)
                    value = cell_info.get("value", "")
                    if value:
                        try:
                            if isinstance(value, str) and value.strip() and not value.startswith("="):
                                num_val = float(value.replace(",", "."))
                                cell.value = num_val
                            else:
                                cell.value = value
                            cells_with_structure_values += 1
                            value_written = True
                        except (ValueError, TypeError):
                            cell.value = str(value)
                            cells_with_structure_values += 1
                            value_written = True
                else:
                    logger.debug(f"[generate_numbers_table_excel] Skipping structure value for formula cell {normalized_cell_addr}")
            
            # If still no value written, leave as empty (None)
            if not value_written:
                cell.value = None
                logger.debug(f"[generate_numbers_table_excel] Empty cell {normalized_cell_addr} (no DB value, no formula, no structure value)")
                cells_empty += 1
            
            # Apply format (prefer format from structure to preserve original formatting)
            # Only use format from cell_map if it's more complete
            structure_format = cell_info.get("format", {})
            updated_format = cell_map.get(normalized_cell_addr, {}).get("format", {}) if isinstance(cell_map.get(normalized_cell_addr), dict) else {}
            # Merge formats: structure format as base, updated format can override
            cell_format = {**structure_format, **updated_format}
            if cell_format:
                # Background color - REMOVE ALL BLACK/DARK COLORS
                if cell_format.get("bg_color"):
                    bg_color = cell_format["bg_color"]
                    if isinstance(bg_color, str):
                        # Remove # if present
                        if bg_color.startswith("#"):
                            bg_color = bg_color[1:]
                        
                        # Normalize to 6-char RGB format for comparison
                        rgb_part = bg_color
                        if len(bg_color) == 8:
                            # ARGB format - extract RGB part (last 6 chars)
                            rgb_part = bg_color[2:].upper()
                        elif len(bg_color) == 6:
                            rgb_part = bg_color.upper()
                        
                        # Check if color is black or very dark (all RGB values <= 5)
                        # Convert hex to decimal to check darkness
                        try:
                            r = int(rgb_part[0:2], 16)
                            g = int(rgb_part[2:4], 16)
                            b = int(rgb_part[4:6], 16)
                            
                            # Skip if all RGB components are <= 5 (very dark/black)
                            if r <= 5 and g <= 5 and b <= 5:
                                logger.info(f"[generate_numbers_table_excel] Skipping black/dark background color {rgb_part} (RGB: {r},{g},{b}) for cell {cell_addr}")
                                # Remove bg_color from format to prevent application
                                cell_format = {k: v for k, v in cell_format.items() if k != "bg_color"}
                            else:
                                # Apply color if not black/dark
                                try:
                                    if len(bg_color) == 6:
                                        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                                        cell.fill = fill
                                    elif len(bg_color) == 8:
                                        # ARGB format - use last 6 chars (RGB)
                                        fill = PatternFill(start_color=bg_color[2:], end_color=bg_color[2:], fill_type="solid")
                                        cell.fill = fill
                                except Exception as fill_error:
                                    logger.warning(f"[generate_numbers_table_excel] Error applying background color {bg_color} to cell {cell_addr}: {fill_error}")
                        except (ValueError, IndexError) as parse_error:
                            logger.warning(f"[generate_numbers_table_excel] Could not parse color {bg_color} for cell {cell_addr}: {parse_error}")
                            # Remove bg_color from format if we can't parse it
                            cell_format = {k: v for k, v in cell_format.items() if k != "bg_color"}
                
                # Font
                font_kwargs = {}
                if cell_format.get("font_color"):
                    font_color = cell_format["font_color"]
                    # Convert to aRGB hex format if needed
                    # openpyxl expects aRGB hex (8 chars: AARRGGBB) or RGB hex (6 chars: RRGGBB)
                    if isinstance(font_color, str):
                        # Remove # if present
                        if font_color.startswith("#"):
                            font_color = font_color[1:]
                        # If it's 6 chars (RGB), it's valid
                        # If it's 8 chars (ARGB), it's valid
                        # If it's something else, try to convert or skip
                        if len(font_color) == 6:
                            # RGB format - add FF for alpha to make it ARGB
                            font_color = "FF" + font_color.upper()
                        elif len(font_color) == 8:
                            # Already ARGB format
                            font_color = font_color.upper()
                        else:
                            # Invalid format - skip color
                            logger.warning(f"[generate_numbers_table_excel] Invalid font_color format: {font_color}, skipping")
                            font_color = None
                        
                        if font_color:
                            font_kwargs["color"] = font_color
                    else:
                        # Not a string - skip
                        logger.warning(f"[generate_numbers_table_excel] font_color is not a string: {type(font_color)}, skipping")
                
                if cell_format.get("bold"):
                    font_kwargs["bold"] = True
                
                if font_kwargs:
                    try:
                        cell.font = Font(**font_kwargs)
                    except Exception as font_error:
                        logger.warning(f"[generate_numbers_table_excel] Error applying font format: {font_error}, skipping font")
                
                # Apply borders if present
                if cell_format.get("borders"):
                    borders_dict = cell_format["borders"]
                    border_style_map = {
                        "thin": Side(style=BORDER_THIN),
                        "medium": Side(style=BORDER_MEDIUM),
                        "thick": Side(style=BORDER_THICK),
                    }
                    border_sides = {}
                    for side_name in ["left", "right", "top", "bottom"]:
                        if borders_dict.get(side_name):
                            style_str = borders_dict[side_name].lower()
                            border_sides[side_name] = border_style_map.get(style_str, Side(style=BORDER_THIN))
                    if border_sides:
                        try:
                            cell.border = Border(**border_sides)
                        except Exception as border_error:
                            logger.warning(f"[generate_numbers_table_excel] Error applying border: {border_error}, skipping")
                
                # Apply alignment if present
                if cell_format.get("alignment"):
                    align_dict = cell_format["alignment"]
                    try:
                        alignment_kwargs = {}
                        if align_dict.get("horizontal"):
                            alignment_kwargs["horizontal"] = align_dict["horizontal"]
                        if align_dict.get("vertical"):
                            alignment_kwargs["vertical"] = align_dict["vertical"]
                        if alignment_kwargs:
                            cell.alignment = Alignment(**alignment_kwargs)
                    except Exception as align_error:
                        logger.warning(f"[generate_numbers_table_excel] Error applying alignment: {align_error}, skipping")
        
        # Set column widths (basic - could be improved by storing original widths)
        max_col = structure.get("columns", 5)
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # CRITICAL: Also write cells that are in DB but NOT in structure (e.g., B5 that was empty in original Excel)
        # These are cells that were added via chat but didn't exist in the original structure
        db_only_cells = set(cell_map.keys()) - {cell_info.get("address", "").upper() for cell_info in structure.get("cells", [])}
        if db_only_cells:
            logger.info(f"[generate_numbers_table_excel] Found {len(db_only_cells)} cells in DB that are not in structure: {sorted(list(db_only_cells))}")
            for cell_addr in db_only_cells:
                cell_data = cell_map[cell_addr]
                value = cell_data.get("value", "") if isinstance(cell_data, dict) else str(cell_data)
                
                # Parse cell address (e.g., "B5" -> row=5, col=2)
                import re
                match = re.match(r'([A-Z]+)(\d+)', cell_addr.upper())
                if match:
                    col_letters = match.group(1)
                    row_num = int(match.group(2))
                    
                    # Convert column letters to number (A=1, B=2, ..., Z=26, AA=27, etc.)
                    col_num = 0
                    for char in col_letters:
                        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                    
                    col_letter = get_column_letter(col_num)
                    cell = ws[f"{col_letter}{row_num}"]
                    
                    # Apply value
                    try:
                        if isinstance(value, str) and value.strip() and not value.startswith("="):
                            num_val = float(value.replace(",", "."))
                            cell.value = num_val
                            logger.info(f"[generate_numbers_table_excel] ✅ Set DB-only cell {cell_addr} to number: {num_val}")
                        else:
                            cell.value = value
                            logger.info(f"[generate_numbers_table_excel] ✅ Set DB-only cell {cell_addr} to value: '{value}'")
                    except (ValueError, TypeError) as e:
                        cell.value = str(value) if value is not None else ""
                        logger.warning(f"[generate_numbers_table_excel] ⚠️ Error converting DB-only cell {cell_addr}: {e}, set to string: '{value}'")
        
        logger.info(f"[generate_numbers_table_excel] Excel generated: {max_col} columns, {structure.get('rows', 0)} rows, {len(structure.get('cells', []))} cells written")
        logger.info(f"[generate_numbers_table_excel] 📊 Summary: {cells_with_db_values} cells with DB values, {cells_with_structure_values} cells with structure values, {cells_with_formulas} cells with formulas, {cells_empty} empty cells, {len(db_only_cells)} DB-only cells added")
        
        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"[generate_numbers_table_excel] Error generating Numbers table Excel: {e}", exc_info=True)
        # DO NOT fallback to old method - re-raise the error
        raise ValueError(f"Error al generar el Excel de la plantilla {template_key}: {str(e)}")


# ------------------ Scenarios & Sensitivity ------------------
def apply_deltas(base: Dict[str, float], deltas: Dict[str, float]) -> Dict[str, float]:
    """Apply multiplicative deltas to base (e.g., {precio_venta: -0.1} means -10%)."""
    out = dict(base)
    for k, pct in (deltas or {}).items():
        v = out.get(k)
        if v is None:
            continue
        try:
            out[k] = float(v) * (1.0 + float(pct))
        except Exception:
            pass
    return out


def what_if(property_id: str, deltas: Dict[str, float], name: str | None = None) -> Dict[str, Any]:
    """Compute a what-if scenario, persist snapshot best-effort, and return outputs.
    Deltas are fractional (e.g., {"precio_venta": -0.1, "costes_construccion": 0.12}).
    """
    rows = get_numbers(property_id)
    base = _to_map(rows)
    scenario_inputs = apply_deltas(base, deltas)
    outputs = compute_derived_from_inputs(scenario_inputs)
    anomalies = validate_anomalies(scenario_inputs, outputs)
    try:
        sb.table("scenario_snapshots").insert({
            "property_id": property_id,
            "name": name or "what_if",
            "deltas": deltas,
            "outputs": outputs,
        }).execute()
    except Exception:
        pass
    return {"inputs": scenario_inputs, "outputs": outputs, "anomalies": anomalies}


def sensitivity_grid(property_id: str, precio_vec: List[float], costes_vec: List[float]) -> Dict[str, Any]:
    """Build a sensitivity grid for net_profit with multiplicative vectors for precio_venta and costes_construccion.
    Vectors contain fractional changes (e.g., [-0.2,-0.1,0,0.1,0.2]).
    """
    rows = get_numbers(property_id)
    base = _to_map(rows)
    grid: List[List[float | None]] = []
    for dv in precio_vec:
        row: List[float | None] = []
        for dc in costes_vec:
            scenario = apply_deltas(base, {"precio_venta": dv, "costes_construccion": dc})
            outs = compute_derived_from_inputs(scenario)
            row.append(outs.get("net_profit"))
        grid.append(row)
    try:
        sb.table("scenario_snapshots").insert({
            "property_id": property_id,
            "name": "sensitivity",
            "deltas": {"precio_vec": precio_vec, "costes_vec": costes_vec},
            "outputs": {"grid": grid},
        }).execute()
    except Exception:
        pass
    return {"precio_vec": precio_vec, "costes_vec": costes_vec, "grid": grid}


def break_even_precio(property_id: str, tol: float = 1.0, max_iter: int = 60) -> Dict[str, Any]:
    """Solve for precio_venta such that net_profit ≈ 0 using bisection on a reasonable bracket.
    Returns {precio_venta, net_profit, iterations} or error.
    """
    rows = get_numbers(property_id)
    base = _to_map(rows)
    # Build a helper to evaluate net_profit for a given precio
    def f(precio: float) -> float | None:
        scenario = dict(base)
        scenario["precio_venta"] = precio
        outs = compute_derived_from_inputs(scenario)
        return outs.get("net_profit")

    # Find a bracket around current precio (or a default)
    p0 = base.get("precio_venta") or 100000.0
    lo = max(1.0, p0 * 0.5)
    hi = p0 * 1.5
    v_lo = f(lo)
    v_hi = f(hi)
    if v_lo is None or v_hi is None:
        return {"error": "insufficient_data"}
    # If same sign, expand a bit
    expand = 0
    while v_lo is not None and v_hi is not None and v_lo * v_hi > 0 and expand < 5:
        lo *= 0.8
        hi *= 1.2
        v_lo = f(lo)
        v_hi = f(hi)
        expand += 1
    if v_lo is None or v_hi is None:
        return {"error": "insufficient_data"}
    # Bisection
    it = 0
    root = None
    while it < max_iter:
        mid = 0.5 * (lo + hi)
        v_mid = f(mid)
        if v_mid is None:
            break
        if abs(v_mid) <= tol:
            root = mid
            break
        # Decide side
        if v_lo * v_mid <= 0:
            hi = mid
            v_hi = v_mid
        else:
            lo = mid
            v_lo = v_mid
        it += 1
    if root is None:
        root = 0.5 * (lo + hi)
    return {"precio_venta": root, "net_profit": f(root), "iterations": it}


# ------------------ Charts (Plotly PNG → Supabase) ------------------
def _save_png(property_id: str, fig, chart_type: str, params: Dict[str, Any] | None = None) -> Dict[str, Any]:
    import time
    import plotly.io as pio
    # Generate PNG bytes with Kaleido
    png_bytes = pio.to_image(fig, format="png", scale=2)
    key = f"charts/{property_id}/{chart_type}/{int(time.time())}.png"
    try:
        sb.storage.from_(BUCKET).upload(key, png_bytes, {"content-type": "image/png", "upsert": "true"})
        signed = sb.storage.from_(BUCKET).create_signed_url(key, 3600)
        # Cache entry (best-effort)
        try:
            sb.table("chart_cache").insert({
                "property_id": property_id,
                "chart_type": chart_type,
                "params": params or {},
                "storage_key": key,
            }).execute()
        except Exception:
            pass
        return {"storage_key": key, "signed_url": signed.get("signedURL")}
    except Exception as e:
        return {"error": str(e)}


def chart_waterfall(property_id: str) -> Dict[str, Any]:
    try:
        import plotly.graph_objects as go
    except ModuleNotFoundError:
        return {"error": "plotly_no_instalado", "hint": "Instala plotly y kaleido (pip install plotly kaleido) y reinicia el servidor."}
    rows = get_numbers(property_id)
    vals = _to_map(rows)
    precio = vals.get("precio_venta")
    buckets = {
        "Project Mgmt": vals.get("project_mgmt_fees"),
        "Terrenos": vals.get("terrenos_coste"),
        "Project Management": vals.get("project_management_coste"),
        "Acometidas": vals.get("acometidas"),
        "Construcción": vals.get("costes_construccion"),
    }
    impuestos_total = None
    if vals.get("impuestos_pct") is not None and precio is not None:
        impuestos_total = vals["impuestos_pct"] * precio

    if precio is None:
        return {"error": "precio_venta requerido"}

    measure = ["relative"] * len(buckets)
    text = []
    y = [-(buckets[k] or 0.0) for k in buckets]
    labels = list(buckets.keys())
    if impuestos_total is not None:
        labels.append("Impuestos")
        y.append(-impuestos_total)
        measure.append("relative")
    # Net profit bar
    # compute derived quickly
    outs = compute_derived_from_inputs(vals)
    net = outs.get("net_profit") or 0.0
    labels.append("Net Profit")
    y.append(net)
    measure.append("total")

    fig = go.Figure(go.Waterfall(
        name="Profit",
        orientation="v",
        measure=measure,
        x=labels,
        textposition="outside",
        y=y,
        decreasing={"marker": {"color": "#c5ac85"}},
        increasing={"marker": {"color": "#6eb55e"}},
        totals={"marker": {"color": "#3d7435"}},
    ))
    fig.update_layout(title="Precio de venta → Net Profit", showlegend=False)
    return _save_png(property_id, fig, "waterfall", {"buckets": list(buckets.keys())})


def chart_cost_stack(property_id: str) -> Dict[str, Any]:
    try:
        import plotly.graph_objects as go
    except ModuleNotFoundError:
        return {"error": "plotly_no_instalado", "hint": "Instala plotly y kaleido (pip install plotly kaleido) y reinicia el servidor."}
    rows = get_numbers(property_id)
    v = _to_map(rows)
    buckets = {
        "Project Mgmt": v.get("project_mgmt_fees"),
        "Terrenos": v.get("terrenos_coste"),
        "Project Management": v.get("project_management_coste"),
        "Acometidas": v.get("acometidas"),
        "Construcción": v.get("costes_construccion"),
    }
    total = sum([x for x in buckets.values() if x is not None]) if any(buckets.values()) else 0.0
    parts = [(k, (buckets[k] or 0.0) / total if total else 0.0) for k in buckets]

    fig = go.Figure()
    fig.add_bar(x=["Composición"], y=[p[1] for p in parts], name="%", marker_color=["#b3dfaa", "#8fcb7f", "#6eb55e", "#4f9542", "#3d7435"] * 2)
    fig.update_layout(barmode="stack", title="Composición de costes (100%)", yaxis=dict(tickformat=",.0%"))
    return _save_png(property_id, fig, "stacked_100", {})


def chart_sensitivity_heatmap(property_id: str, precio_vec: List[float], costes_vec: List[float]) -> Dict[str, Any]:
    try:
        import plotly.graph_objects as go
    except ModuleNotFoundError:
        return {"error": "plotly_no_instalado", "hint": "Instala plotly y kaleido (pip install plotly kaleido) y reinicia el servidor."}
    grid = sensitivity_grid(property_id, precio_vec, costes_vec)
    z = grid.get("grid") or []
    # UI palette: earth → neutral → green
    campo_colorscale = [
        [0.0, "#c5ac85"],   # earth (negative)
        [0.5, "#d4eece"],   # neutral light green
        [1.0, "#3d7435"],   # deep green (positive)
    ]
    fig = go.Figure(data=go.Heatmap(z=z, x=costes_vec, y=precio_vec, colorscale=campo_colorscale))
    fig.update_layout(title="Sensibilidad net_profit (precio vs construcción)", xaxis_title="Δ costes_construccion", yaxis_title="Δ precio_venta")
    return _save_png(property_id, fig, "sensitivity_heatmap", {"precio_vec": precio_vec, "costes_vec": costes_vec})




